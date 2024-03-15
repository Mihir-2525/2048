import numpy as np
import random
import keyboard
import os
import time
import xlsxwriter
import openpyxl

# Return Check to user get 2048 Number  
def win(arr):
    for i in arr:
        for j in i:
            if j==2048:return True

# Check if the array is empty or not
def check(arr):
    # np.all() returns True if the array doesn't have any 0 value in any element
    return not np.all(arr)

# Horizontally check for duplicates
def check_Horizontal(x):
    for i in x:
        for j in range(len(i)-1):
            if i[j] == i[j+1]:
                return True

# Vertically check for duplicates
def check_Vertical(x):
    for i in range(len(x)-1):
        for j in range(len(x)):
            if x[i,j] == x[i+1,j]:
                return True

# Add a new element to the array
def new_var(arr, size, difficulty):
    # Check if the array doesn't contain zero, then don't add any new elements
    while check(arr):
        # Set random dimensions
        x, y = random.randint(0, size-1), random.randint(0, size-1)
        if arr[x, y] == 0:
            if difficulty == 1:
                arr[x, y] = random.choice([2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 4, 4])
            elif difficulty == 2:
                arr[x, y] = random.choice([2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 2, 4, 4, 4, 8])
            elif difficulty == 3:
                arr[x, y] = random.choice([2, 2, 2, 2, 2, 2, 2, 2, 2, 4, 4, 4, 8])
            elif difficulty == 2048:
                arr[x, y] = random.choice([128,128,256])
            break

# <---
def right_to_left(arr, size):
    for i in range(size):
        # shift all non-zero elements to the left
        non_zero_elements = [num for num in arr[i] if num != 0]
        # merge elements if they are the same and next to each other
        merged_elements = []
        skip = False
        for j in range(len(non_zero_elements)):
            if skip:
                skip = False
                continue
            if j < len(non_zero_elements) - 1 and non_zero_elements[j] == non_zero_elements[j + 1]:
                merged_elements.append(non_zero_elements[j] * 2)
                skip = True
            else:
                merged_elements.append(non_zero_elements[j])
        # Fill the remaining spaces with zeros
        merged_elements.extend([0] * (size - len(merged_elements)))
        arr[i] = merged_elements

# --->
def left_to_right(arr, size):
    for i in range(size):
        # shift all non-zero elements to the right
        non_zero_elements = [num for num in arr[i] if num != 0]
        # merge elements if they are the same and next to each other
        merged_elements = []
        skip = False
        for j in range(len(non_zero_elements) - 1, -1, -1):
            if skip:
                skip = False
                continue
            if j > 0 and non_zero_elements[j] == non_zero_elements[j - 1]:
                merged_elements.append(non_zero_elements[j] * 2)
                skip = True
            else:
                merged_elements.append(non_zero_elements[j])
        # Fill the remaining spaces with zeros
        merged_elements = [0] * (size - len(merged_elements)) + merged_elements[::-1]
        arr[i] = merged_elements

# ^
# |
def bottom_to_top(arr, size):
    for j in range(size):
        # shift all non-zero elements upwards
        non_zero_elements = [arr[i][j] for i in range(size) if arr[i][j] != 0]
        # merge elements if they are the same and next to each other
        merged_elements = []
        skip = False
        for i in range(len(non_zero_elements)):
            if skip:
                skip = False
                continue
            if i < len(non_zero_elements) - 1 and non_zero_elements[i] == non_zero_elements[i + 1]:
                merged_elements.append(non_zero_elements[i] * 2)
                skip = True
            else:
                merged_elements.append(non_zero_elements[i])
        # Fill the remaining spaces with zeros
        merged_elements = merged_elements + [0] * (size - len(merged_elements))
        for i in range(size):
            arr[i][j] = merged_elements[i]

# |
# V
def top_to_bottom(arr, size):
    for j in range(size):
        # shift all non-zero elements downwards
        non_zero_elements = [arr[i][j] for i in range(size) if arr[i][j] != 0]
        # merge elements if they are the same and next to each other
        merged_elements = []
        skip = False
        for i in range(len(non_zero_elements) - 1, -1, -1):
            if skip:
                skip = False
                continue
            if i > 0 and non_zero_elements[i] == non_zero_elements[i - 1]:
                merged_elements.append(non_zero_elements[i] * 2)
                skip = True
            else:
                merged_elements.append(non_zero_elements[i])
        # Fill the remaining spaces with zeros
        merged_elements = [0] * (size - len(merged_elements)) + merged_elements[::-1]
        for i in range(size):
            arr[i][j] = merged_elements[i]

# Create Excel file to Store the data
def create_file(filename):
    # Check if file exists, if not, create it
    if not os.path.exists(filename):
        workbook = xlsxwriter.Workbook(filename)
        workbook.close()

# set default data into the workbook
def default_workbook(worksheet,):
    worksheet['B1'] = 'Difficulty Level'
    for i in range(2,17):worksheet['B'+str(i)] = 'Easy'
    worksheet['C1'] = 'Highest Score'
    worksheet['D1'] = 'Difficulty Level'
    for i in range(2,17):worksheet['D'+str(i)] = 'Medium'
    worksheet['E1'] = 'Highest Score'
    worksheet['F1'] = 'Difficulty Level'
    for i in range(2,17):worksheet['F'+str(i)] = 'Hard'
    worksheet['G1'] = 'Highest Score'
    worksheet['I1'] = 'Last Played Level'
    col = 'A'
    row = 2
    for i in range(15):
        # Write data to the worksheet
        # A
        worksheet[col+str(row)] = str(row)+"x"+str(row)
        # if data not stored in worksheet then write 0 
        # C
        if (worksheet[chr(ord(col) + 2)+str(row)].value) == None:
            worksheet[chr(ord(col) + 2)+str(row)] = 0
        # E
        if (worksheet[chr(ord(col) + 4)+str(row)].value) == None:
            worksheet[chr(ord(col) + 4)+str(row)] = 0
        # G
        if (worksheet[chr(ord(col) + 6)+str(row)].value) == None:
            worksheet[chr(ord(col) + 6)+str(row)] = 0
        row = row + 1

def Column_difficulty(difficulty):
    if difficulty == 1:
        return 'G'
    elif difficulty == 2:
        return 'E'
    elif difficulty == 3:
        return 'C'
    elif difficulty == 2048:
        return 'H'

def max_arr(arr):
    max_n = 0
    for i in arr:
        for j in i:
            if j > max_n: max_n = j
    return max_n

def add_data(arr, size, difficulty, filename, workbook, worksheet):    
    # Find Max number from the list
    max_n = max_arr(arr)
    
    # Set in which Column we add Score 
    col = Column_difficulty(difficulty)
    
    # Write the data to the file
    try:
        if max_n > worksheet[col+str(size)].value:
            worksheet[col+str(size)] = max_n
            workbook.save(filename)
    except Exception:
        worksheet[col+str(size)] = max_n
        workbook.save(filename)

    # Open the workbook again to read the data
    workbook = openpyxl.load_workbook(filename)

    # Get the active worksheet
    worksheet = workbook.active

# Print data on console display
def print_board(arr, size, difficulty, worksheet):
    # Clear the screen
    os.system("cls")
    
    # Display in which Column we Contain Score 
    col = Column_difficulty(difficulty)

    # Find max Number from List
    max_num = max_arr(arr)
    highscore = worksheet[col+str(size)].value
    try:
        if(highscore < max_num):highscore = max_num
    except Exception:
        highscore = max_num

    # Read the value from cell
    print("Highest Score is ",highscore)
    print("Current Score is ",max_num)

    # Print the game board
    # ╔═══════════════╦═══════════════╦═══════════════╦═══════════════╗
    print("╔"+"═"*15,end="")
    for i in range(len(arr)-1):
        print("╦"+"═"*15,end="")
    print("╗")
    # Print single line
    for i in range(len(arr)):
        # ║               ║               ║               ║               ║
        print("║",end="")
        for j in range(len(arr)):
            print("\t\t║",end="")
        print()
        # ║       x       ║       x       ║       x       ║       x       ║
        for j in range(len(arr)):
            print("║\t",end="")
            print('',end="") if arr[i][j] == 0 else print(str(arr[i][j]),end="")
            print("\t",end="")
        print("║")
        # ║               ║               ║               ║               ║
        print("║",end="")
        for j in range(len(arr)):
            print("\t\t║",end="")
        print()
        # ╠═══════════════╬═══════════════╬═══════════════╬═══════════════╣
        if not(i==len(arr)-1):
            print("╠"+"═"*15,end="")
            for i in range(len(arr)-1):
                print("╬"+"═"*15,end="")
            print("╣")
    # ╚═══════════════╩═══════════════╩═══════════════╩═══════════════╝
    print("╚"+"═"*15,end="")
    for i in range(len(arr)-1):
        print("╩"+"═"*15,end="")
    print("╝")

def main():
    # Create File to Store Data
    filename = 'data.xlsx'
    
    # Create Excel file to Store the data
    create_file(filename)

    # Open the existing or newly created workbook
    workbook = openpyxl.load_workbook(filename)

    # Get the active worksheet
    worksheet = workbook.active
    
    # set default data into the workbook
    default_workbook(worksheet)

    # Save the workbook
    workbook.save(filename)
    
    # Clear the screen
    os.system("cls")

    # Set size of the array
    size = 0
    while size < 2 or size > 17:
        try:
            size = int(input("Enter the number of the box you want: "))
        except ValueError:
            print("Invalid input. Please enter a valid Size.")
    # Set Difficulty
    difficulty = 0
    while difficulty not in [1, 2, 3, 2048]:
        try:
            difficulty = int(input("\n1 : Hard\n2 : Medium\n3 : Easy\nEnter the Level of difficulty of the game: "))
            if difficulty in [1, 2, 3, 2048]:
                break
            else:
                print("Invalid input. Please enter 1, 2, or 3.")
        except ValueError:
            print("Invalid input. Please enter a valid number.")
    print(f"Selected difficulty: {difficulty}")
    # B
    for i in range(15):
        if difficulty == 1: d = 'Hard'
        elif difficulty == 2: d = 'Medium'
        elif difficulty == 3: d = 'Easy'
        else: d = 'Special'
        # Store Last Played Game's Difficulty Level in Excel
        worksheet['I'+str(size)].value = d
        workbook.save(filename)

    # Set the array (i.e., if size = 3, then 'arr' makes a 3x3 empty (0) array)
    arr = np.zeros(size*size, dtype=np.int32).reshape(size, size)

    new_var(arr, size,difficulty)
    print_board(arr, size, difficulty, worksheet)

    while True:
        if (check(arr) or check_Horizontal(arr) or check_Vertical(arr)):
            # Add data into Excel file
            add_data(arr, size, difficulty, filename, workbook, worksheet)
            if win(arr):
                print_board(arr, size, difficulty, worksheet)
                print("╔═══╗╔═══╗╔═╗─╔╗╔═══╗╔═══╗╔═══╗╔════╗╔╗─╔╗╔╗───╔═══╗╔════╗╔══╗╔═══╗╔═╗─╔╗╔═══╗")
                print("║╔═╗║║╔═╗║║║╚╗║║║╔═╗║║╔═╗║║╔═╗║║╔╗╔╗║║║─║║║║───║╔═╗║║╔╗╔╗║╚╣─╝║╔═╗║║║╚╗║║║╔═╗║")
                print("║║─╚╝║║─║║║╔╗╚╝║║║─╚╝║╚═╝║║║─║║╚╝║║╚╝║║─║║║║───║║─║║╚╝║║╚╝─║║─║║─║║║╔╗╚╝║║╚══╗")
                print("║║─╔╗║║─║║║║╚╗║║║║╔═╗║╔╗╔╝║╚═╝║──║║──║║─║║║║─╔╗║╚═╝║──║║───║║─║║─║║║║╚╗║║╚══╗║")
                print("║╚═╝║║╚═╝║║║─║║║║╚╩═║║║║╚╗║╔═╗║──║║──║╚═╝║║╚═╝║║╔═╗║──║║──╔╣─╗║╚═╝║║║─║║║║╚═╝║")
                print("╚═══╝╚═══╝╚╝─╚═╝╚═══╝╚╝╚═╝╚╝─╚╝──╚╝──╚═══╝╚═══╝╚╝─╚╝──╚╝──╚══╝╚═══╝╚╝─╚═╝╚═══╝")
                break
            
            key = keyboard.read_key().lower()
            if key == "w" or key == "up":
                bottom_to_top(arr, size)
            elif key == "s" or key == "down":
                top_to_bottom(arr, size)
            elif key == "a" or key == "left":
                right_to_left(arr, size)
            elif key == "d" or key == "right":
                left_to_right(arr, size)
            elif key == "x":
                break
            else:
                print('Invalid key')
                continue
            # Set delay for the next move
            time.sleep(0.2)
            # Create a new variable
            new_var(arr, size, difficulty)
            # Print the updated box
            print_board(arr, size, difficulty, worksheet)
        else:
            print("╔╗──╔╗╔═══╗╔╗─╔╗   ╔╗───╔═══╗╔═══╗╔═══╗   ╔════╗╔╗─╔╗╔═══╗   ╔═══╗╔═══╗╔═╗╔═╗╔═══╗")
            print("║╚╗╔╝║║╔═╗║║║─║║   ║║───║╔═╗║║╔═╗║║╔══╝   ║╔╗╔╗║║║─║║║╔══╝   ║╔═╗║║╔═╗║║║╚╝║║║╔══╝")
            print("╚╗╚╝╔╝║║─║║║║─║║   ║║───║║─║║║╚══╗║╚══╗   ╚╝║║╚╝║╚═╝║║╚══╗   ║║─╚╝║║─║║║╔╗╔╗║║╚══╗")
            print("─╚╗╔╝─║║─║║║║─║║   ║║─╔╗║║─║║╚══╗║║╔══╝   ──║║──║╔═╗║║╔══╝   ║║╔═╗║╚═╝║║║║║║║║╔══╝")
            print("──║║──║╚═╝║║╚═╝║   ║╚═╝║║╚═╝║║╚═╝║║╚══╗   ──║║──║║─║║║╚══╗   ║╚╩═║║╔═╗║║║║║║║║╚══╗")
            print("──╚╝──╚═══╝╚═══╝   ╚═══╝╚═══╝╚═══╝╚═══╝   ──╚╝──╚╝─╚╝╚═══╝   ╚═══╝╚╝─╚╝╚╝╚╝╚╝╚═══╝")
            break

if __name__ == "__main__":
    main()